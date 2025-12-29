from pathlib import Path
from openpyxl import Workbook


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
sheet_name = 'Minha planilha'
workbook.create_sheet(sheet_name, 0)
worksheet = workbook[sheet_name]

workbook.remove(workbook['Sheet'])

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')


students = [
    # nome     idade  nota
    ['João',     14,   5.5],
    ['Maria',    13,   9.7],
    ['Luiz',     16,   8.8],
    ['Alberto',  15,   10 ],
]

# for i, studentsrow in enumerate(students, start=2):
#     for j, studentscollum in enumerate(studentsrow, start=1):
#         worksheet.cell(i, j, studentscollum)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)